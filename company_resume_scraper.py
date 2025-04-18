import asyncio
import time
import os
import logging
import pandas as pd
import json
from datetime import datetime
from playwright.async_api import async_playwright
import tempfile
import shutil
import re
from urllib.parse import quote, unquote
from playwright.async_api import TimeoutError
import traceback
from io import BytesIO
import base64
import requests
from PIL import Image
from job_scraper_final import clean_text_for_excel

# 設置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("resume_scraper.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("104_resume_scraper")

class ResumeScraperConfig:
    """爬蟲配置類"""
    def __init__(self, username="", password="", search_keyword="", page_limit=1):
        self.username = username  # 104企業會員帳號
        self.password = password  # 104企業會員密碼
        self.search_keyword = search_keyword  # 搜尋關鍵字
        self.page_limit = page_limit  # 爬取頁數限制
        
        # 網站URL
        self.vip_url = "https://vip.104.com.tw/"  # VIP系統首頁
        self.search_url = "https://vip.104.com.tw/search"  # 搜尋頁面URL
        
        # 固定的狀態保存目錄（避免每次運行都建立新目錄）
        self.data_dir = "104_data"
        os.makedirs(self.data_dir, exist_ok=True)
        
        # 每次運行的結果目錄
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_dir = f"resume_data_{self.timestamp}"
        os.makedirs(self.output_dir, exist_ok=True)

class ResumeScraper:
    """104求職者資料爬蟲類"""
    
    def __init__(self, config=None):
        self.config = config or ResumeScraperConfig()
        self.browser = None
        self.context = None
        self.page = None
    
    async def initialize(self):
        """初始化瀏覽器，使用持久化上下文保存登入狀態"""
        playwright = await async_playwright().start()
        
        # 建立用戶資料目錄
        user_data_dir = os.path.join(self.config.data_dir, "browser_data")
        os.makedirs(user_data_dir, exist_ok=True)
        
        # 使用持久化上下文替代傳統的 launch + new_context 方式
        self.browser = await playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=False,
            accept_downloads=True,
            bypass_csp=True,
            slow_mo=500,  # 適當放慢操作速度，避免被反爬
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-web-security',
                '--disable-features=IsolateOrigins,site-per-process'
            ],
            viewport={"width": 1280, "height": 800},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="zh-TW",
            timezone_id="Asia/Taipei",
            extra_http_headers={
                "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8"
            }
        )
        
        # 修改WebDriver相關屬性，避免被檢測
        await self.browser.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => false,
            });
            window.chrome = { runtime: {} };
        """)
        
        # 在持久化上下文中創建頁面
        self.page = await self.browser.new_page()
        logger.info("瀏覽器初始化成功，使用持久化上下文模式")
    
    async def login(self):
        """改進的登入流程，利用持久化上下文自動管理登入狀態"""
        try:
            # 步驟1: 進入VIP首頁
            await self.page.goto(self.config.vip_url)
            logger.info("已進入VIP首頁")
            
            # 等待頁面加載
            await self.page.wait_for_load_state('networkidle', timeout=15000)
            
            # 檢查是否已經登入
            is_logged_in = await self.check_if_logged_in()
            if is_logged_in:
                logger.info("✅ 檢測到您已經登入，無需重新驗證!")
                return True
            
            # 步驟2: 檢查是否已經在repeatLogin頁面
            current_url = self.page.url
            if "repeatLogin" in current_url:
                logger.info("檢測到重複登入頁面")
                
                # 等待頁面完全載入
                await self.page.wait_for_load_state('networkidle', timeout=15000)
                
                # 找到並點擊"將目前帳號登出，立即登入"按鈕
                try:
                    # 先保存頁面截圖來分析問題
                    screenshot_path = os.path.join(self.config.output_dir, f"repeat_login_{int(time.time())}.png")
                    await self.page.screenshot(path=screenshot_path)
                    
                    # 嘗試列出頁面上所有按鈕及其文本以便診斷
                    button_texts = await self.page.evaluate('''() => {
                        return Array.from(document.querySelectorAll('button')).map(b => b.textContent.trim());
                    }''')
                    logger.info(f"頁面上所有按鈕文本: {button_texts}")
                    
                    # 使用更嚴格的選擇器策略
                    await self.page.click('button:has-text("將目前帳號登出，立即登入")', strict=True, timeout=5000)
                    logger.info("成功點擊「將目前帳號登出，立即登入」按鈕")
                except Exception as e:
                    logger.warning(f"嚴格選擇器失敗: {e}")
                    
                    # 嘗試使用最簡單的文本選擇方式
                    try:
                        await self.page.click('text="將目前帳號登出，立即登入"')
                        logger.info("通過文本內容點擊按鈕成功")
                    except Exception as text_e:
                        logger.warning(f"文本選擇器失敗: {text_e}")
                        
                        # 最後嘗試通過JavaScript找出並點擊按鈕
                        try:
                            result = await self.page.evaluate('''() => {
                                // 嘗試找到精確包含特定文本的按鈕
                                const allElements = document.querySelectorAll('*');
                                for (const el of allElements) {
                                    if (el.textContent.includes('將目前帳號登出，立即登入')) {
                                        el.click();
                                        return `點擊了包含文本的元素: ${el.tagName}`;
                                    }
                                }
                                return "未找到按鈕";
                            }''')
                            logger.info(f"JavaScript執行結果: {result}")
                        except Exception as js_e:
                            logger.error(f"JavaScript方法失敗: {js_e}")
                
                # 等待更長時間確保頁面轉換
                await asyncio.sleep(8)
                
                # 檢查是否仍在重複登入頁面
                if "repeatLogin" in self.page.url:
                    logger.warning("仍在重複登入頁面，嘗試直接訪問首頁")
                    await self.page.goto("https://vip.104.com.tw/index/index")
                    await asyncio.sleep(5)
            
            # 再次檢查是否已經登入
            is_logged_in = await self.check_if_logged_in()
            if is_logged_in:
                logger.info("✅ 通過重複登入處理後已成功登入!")
                return True
            
            # 步驟3: 點擊登入按鈕
            login_button_selectors = [
                'a:text("登入")', 
                'button:text("登入")', 
                '.login-btn',
                'a[href*="login"]'
            ]
            
            for selector in login_button_selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    await self.page.click(selector)
                    logger.info(f"已點擊登入按鈕 '{selector}'")
                    break
                except Exception as e:
                    logger.debug(f"點擊登入按鈕 '{selector}' 失敗: {e}")
                    continue
            
            # 步驟4: 等待登入表單加載
            await asyncio.sleep(2)
            
            # 步驟5: 填寫帳號
            await self.page.fill('input[type="text"]', self.config.username)
            logger.info("已填寫帳號")
            
            # 步驟6: 填寫密碼
            await self.page.fill('input[type="password"]', self.config.password)
            logger.info("已填寫密碼")
            
            # 步驟7: 點擊登入按鈕
            submit_button_selectors = [
                'button[type="submit"]',
                'button:text("登入")',
                'input[type="submit"]'
            ]
            
            for selector in submit_button_selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    await self.page.click(selector)
                    logger.info(f"已點擊提交按鈕 '{selector}'")
                    break
                except Exception as e:
                    logger.debug(f"點擊提交按鈕 '{selector}' 失敗: {e}")
                    continue
            
            # 步驟8: 等待系統處理，準備接收驗證碼
            logger.info("登入提交完成，準備接收郵箱驗證碼...")
            
            # 步驟9: 等待驗證碼頁面加載
            await asyncio.sleep(5)
            
            # 步驟10: 檢查當前頁面，判斷是否已經成功登入或需要驗證碼
            current_url = self.page.url
            if "/index/index" in current_url:
                logger.info("已直接成功登入VIP系統，無需輸入驗證碼")
                return True
            
            # 步驟11: 等待用戶輸入驗證碼
            logger.info("請等待3秒，系統正在發送驗證碼到您的信箱...")
            for i in range(3, 0, -5):
                logger.info(f"剩餘等待時間: {i} 秒")
                await asyncio.sleep(5)
            
            logger.info("請檢查您的電子郵件並輸入收到的驗證碼")
            user_captcha = input("請輸入您收到的郵箱驗證碼: ")
            
            # 步驟12: 填入驗證碼
            captcha_selectors = [
                'input[name="captcha"]',
                'input[placeholder*="驗證碼"]',
                'input[type="text"]:not([autocomplete="username"])'
            ]
            
            for selector in captcha_selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    await self.page.fill(selector, user_captcha)
                    logger.info(f"已填入驗證碼")
                    break
                except Exception as e:
                    logger.debug(f"填入驗證碼失敗: {e}")
                    continue
            
            # 步驟13: 提交驗證碼
            submit_captcha_selectors = [
                'button[type="submit"]',
                'button:text("確認")',
                'button:text("提交")',
                'input[type="submit"]'
            ]
            
            for selector in submit_captcha_selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    await self.page.click(selector)
                    logger.info(f"已提交驗證碼")
                    break
                except Exception as e:
                    logger.debug(f"提交驗證碼失敗: {e}")
                    continue
            
            # 步驟14: 等待驗證處理
            await asyncio.sleep(5)
            
            # 步驟15: 最終檢查是否登入成功
            current_url = self.page.url
            if "/index/index" in current_url:
                logger.info("驗證成功，已登入VIP系統")
                return True
            else:
                logger.warning(f"可能未成功登入，當前URL: {current_url}")
                return False
            
        except Exception as e:
            logger.error(f"登入過程發生異常: {e}")
            screenshot_path = os.path.join(self.config.output_dir, f"login_error_{int(time.time())}.png")
            await self.page.screenshot(path=screenshot_path)
            logger.info(f"錯誤頁面截圖已保存至: {screenshot_path}")
            return False
    
    async def search(self):
        """搜索履歷功能"""
        try:
            # 檢查是否在首頁
            current_url = self.page.url
            if "/index/index" not in current_url and "vip.104.com.tw" not in current_url:
                logger.warning("似乎未在VIP系統主頁，嘗試重新導航")
                await self.page.goto(self.config.vip_url)
                await asyncio.sleep(3)
            
            logger.info("開始搜尋流程")
            
            # 嘗試找到搜尋輸入框
            search_input_selectors = [
                'input[data-qa-id="inputKeywordSearch"]',
                'input[placeholder="請輸入關鍵字"]',
                'input.form-input--dark',
                'input.form-control.form-input',
                'input[placeholder*="搜尋"]',
                'input[type="search"]',
                '.search-input',
                'input[name="keyword"]',
                'input[data-testid="search-input"]',
                '.header-search input'
            ]
            
            search_input_found = False
            for selector in search_input_selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    await self.page.fill(selector, self.config.search_keyword)
                    logger.info(f"已輸入搜尋關鍵詞: {self.config.search_keyword}")
                    search_input_found = True
                    
                    # 按下Enter鍵執行搜尋
                    await self.page.press(selector, 'Enter')
                    logger.info("已按下Enter鍵執行搜尋")
                    break
                except Exception as e:
                    logger.debug(f"使用選擇器 '{selector}' 搜尋失敗: {e}")
                    continue
            
            if not search_input_found:
                logger.warning("無法找到搜尋輸入框，嘗試直接進入搜尋頁面")
                await self.page.goto(self.config.search_url)
                await asyncio.sleep(3)
                
                # 再次嘗試在搜尋頁面找到輸入框
                for selector in search_input_selectors:
                    try:
                        await self.page.wait_for_selector(selector, timeout=5000)
                        await self.page.fill(selector, self.config.search_keyword)
                        logger.info(f"已在搜尋頁面輸入關鍵詞: {self.config.search_keyword}")
                        
                        # 按下Enter鍵執行搜尋
                        await self.page.press(selector, 'Enter')
                        logger.info("已按下Enter鍵執行搜尋")
                        search_input_found = True
                        break
                    except Exception as e:
                        logger.debug(f"在搜尋頁面使用選擇器 '{selector}' 失敗: {e}")
                        continue
            
            if not search_input_found:
                logger.error("無法找到搜尋輸入框，搜尋失敗")
                return False
            
            # 等待搜尋結果加載
            logger.info("等待搜尋結果加載...")
            await asyncio.sleep(5)
            
            # 確認搜尋成功
            current_url = self.page.url
            if "search" in current_url:
                logger.info(f"搜尋成功，當前URL: {current_url}")
                
                # 截圖搜尋結果頁面
                screenshot_path = os.path.join(self.config.output_dir, f"search_results_{int(time.time())}.png")
                await self.page.screenshot(path=screenshot_path)
                logger.info(f"搜尋結果頁面截圖已保存至: {screenshot_path}")
                
                return True
            else:
                logger.warning(f"搜尋可能未成功，當前URL: {current_url}")
                return False
            
        except Exception as e:
            logger.error(f"搜尋過程發生異常: {e}")
            screenshot_path = os.path.join(self.config.output_dir, f"search_error_{int(time.time())}.png")
            await self.page.screenshot(path=screenshot_path)
            logger.info(f"錯誤頁面截圖已保存至: {screenshot_path}")
            return False
    
    async def extract_results(self):
        """從搜尋結果頁面提取履歷卡片信息，支持多頁提取"""
        try:
            logger.info("開始提取搜尋頁面的履歷卡片信息")
            
            # 建立照片存儲目錄
            photos_dir = os.path.join(self.config.output_dir, "profile_photos")
            os.makedirs(photos_dir, exist_ok=True)
            
            # 儲存所有頁面的履歷卡片和照片
            all_resume_cards = []
            all_photo_files = []
            
            current_page = 1
            
            while current_page <= self.config.page_limit:
                logger.info(f"正在處理第 {current_page}/{self.config.page_limit} 頁")
                
                # 儲存當前頁面的截圖
                screenshot_path = os.path.join(self.config.output_dir, f"page_{current_page}_{int(time.time())}.png")
                await self.page.screenshot(path=screenshot_path)
                
                # 嘗試使用多種選擇器找到履歷卡片
                card_selectors = [
                    '.resume-card',
                    '.candidate-card',
                    '.search-result-item',
                    '.list-item',
                    '[role="listitem"]',
                    '.card',
                    'article',
                    '.resumeList',
                    '.BaseCard'  # 104常用的卡片容器
                ]
                
                resume_cards = []
                photo_files = []  # 儲存照片文件路徑和對應的索引
                
                # 處理卡片
                for selector in card_selectors:
                    try:
                        cards = await self.page.query_selector_all(selector)
                        if cards and len(cards) > 0:
                            logger.info(f"找到 {len(cards)} 個履歷卡片，使用選擇器: {selector}")
                            
                            # 從每個卡片中提取信息
                            for i, card in enumerate(cards):
                                try:
                                    # 獲取卡片的HTML和文本內容
                                    card_html = await self.page.evaluate('(element) => element.outerHTML', card)
                                    card_text = await self.page.evaluate('(element) => element.textContent', card)
                                    
                                    # 使用字串比對從文本中提取各種信息
                                    resume_info = self.extract_info_from_text(card_text)
                                    
                                    # 增加頁碼信息
                                    resume_info['page_number'] = current_page
                                    
                                    # 獲取大頭照URL
                                    photo_url = await self.extract_photo_url(card)
                                    resume_info['photo_url'] = photo_url
                                    
                                    # 下載照片
                                    if photo_url:
                                        try:
                                            # 使用姓名(或索引)作為文件名的一部分
                                            name_value = resume_info.get('name')
                                            if name_value:
                                                safe_name = name_value.replace(' ', '_')
                                            else:
                                                safe_name = f'person_p{current_page}_{i}'
                                            
                                            filename = f"{safe_name}_{int(time.time())}.jpg"
                                            filename = self.sanitize_filename(filename)
                                            photo_path = os.path.join(photos_dir, filename)
                                            
                                            # 下載照片
                                            success = await self.download_photo(photo_url, photo_path)
                                            if success:
                                                resume_info['photo_path'] = photo_path
                                                # 記錄目前照片的索引，須考慮已處理的卡片數量
                                                photo_idx = len(all_resume_cards) + len(resume_cards) + 1
                                                photo_files.append((photo_idx, photo_path))
                                            else:
                                                logger.warning(f"下載照片失敗: {photo_url}")
                                            
                                        except Exception as photo_error:
                                            logger.error(f"處理大頭照過程中發生錯誤: {photo_error}")
                                    
                                    resume_cards.append(resume_info)
                                    logger.info(f"已提取第 {current_page} 頁第 {i+1} 個履歷卡片: {resume_info.get('name', '未知姓名')}")
                                    
                                except Exception as e:
                                    logger.error(f"提取第 {current_page} 頁第 {i+1} 個履歷卡片時出錯: {e}")
                            
                            break  # 找到並處理卡片後退出循環
                    except Exception as e:
                        logger.debug(f"使用選擇器 '{selector}' 查找卡片時發生錯誤: {e}")
                        continue
                
                if resume_cards:
                    # 將當前頁的卡片信息添加到總結果中
                    all_resume_cards.extend(resume_cards)
                    all_photo_files.extend(photo_files)
                    
                    # 保存當前進度到臨時文件
                    temp_df = pd.DataFrame(all_resume_cards)
                    
                    # 清理數據以防止Excel錯誤
                    for column in temp_df.columns:
                        if temp_df[column].dtype == 'object':  # 只處理字符串類型的列
                            temp_df[column] = temp_df[column].apply(lambda x: clean_text_for_excel(x) if isinstance(x, str) else x)
                    
                    temp_path = os.path.join(self.config.output_dir, f"履歷資料_進度_到第{current_page}頁.xlsx")
                    temp_df.to_excel(temp_path, index=False, engine='openpyxl')
                    logger.info(f"已保存當前進度至: {temp_path}")
                    
                    # 檢查是否需要繼續提取下一頁
                    if current_page < self.config.page_limit:
                        has_next = await self.go_to_next_page()
                        if has_next:
                            current_page += 1
                            await asyncio.sleep(3)  # 等待更長時間確保頁面已完全加載
                        else:
                            logger.info("沒有下一頁或進入下一頁失敗，停止提取")
                            break
                    else:
                        logger.info(f"已達到設置的頁數限制 ({self.config.page_limit} 頁)，停止提取")
                        break
                else:
                    logger.warning(f"第 {current_page} 頁未找到任何履歷卡片，停止提取")
                    break
            
            # 所有頁面處理完畢，保存最終結果
            if all_resume_cards:
                # 創建DataFrame
                df = pd.DataFrame(all_resume_cards)
                
                # 清理數據以防止Excel錯誤
                for column in df.columns:
                    if df[column].dtype == 'object':  # 只處理字符串類型的列
                        df[column] = df[column].apply(lambda x: clean_text_for_excel(x) if isinstance(x, str) else x)
                
                # 保存Excel並插入照片
                excel_path = os.path.join(self.config.output_dir, f"履歷資料_共{current_page}頁_{int(time.time())}.xlsx")
                
                # 現有的Excel保存和照片插入邏輯，只需修改照片的索引計算
                try:
                    # 使用openpyxl保存Excel並插入圖片
                    import openpyxl
                    from openpyxl.drawing.image import Image
                    from PIL import Image as PILImage
                    import io
                    
                    # 先將dataframe保存為Excel
                    df.to_excel(excel_path, index=False, engine='openpyxl')
                    
                    # 打開Excel文件以插入圖片
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    
                    # 找到並處理圖片列
                    # 找到 "photo_url" 列的索引
                    photo_col = None
                    for col_idx, cell in enumerate(ws[1], start=1):
                        if cell.value == "photo_url" or cell.value == "photo_path":
                            photo_col = col_idx
                            break
                    
                    if photo_col:
                        # 增加列寬以便更好地顯示圖片
                        ws.column_dimensions[openpyxl.utils.get_column_letter(photo_col)].width = 20
                    
                    # 插入照片
                    for row_idx, photo_path in all_photo_files:
                        try:
                            logger.info(f"處理照片: {photo_path}")
                            if os.path.exists(photo_path) and os.path.getsize(photo_path) > 100:
                                # 使用內存方式處理圖片，避免臨時文件問題
                                try:
                                    # 讀取圖片到內存中
                                    with open(photo_path, 'rb') as img_file:
                                        img_data = img_file.read()
                                    
                                    # 使用BytesIO處理圖片縮放
                                    img_io = io.BytesIO(img_data)
                                    img = PILImage.open(img_io)
                                    
                                    # 調整大小，保持縱橫比
                                    max_size = (100, 100)
                                    img.thumbnail(max_size)
                                    
                                    # 將調整後的圖片保存到內存
                                    output_io = io.BytesIO()
                                    img_format = img.format if img.format else 'JPEG'
                                    img.save(output_io, format=img_format)
                                    output_io.seek(0)
                                    
                                    # 創建Excel圖片對象並插入
                                    if photo_col:
                                        # 使用內存中的圖片創建Excel圖片對象
                                        excel_img = Image(output_io)
                                        
                                        # 插入圖片
                                        cell = ws.cell(row=row_idx+1, column=photo_col)
                                        ws.add_image(excel_img, cell.coordinate)
                                        
                                        # 調整行高
                                        ws.row_dimensions[row_idx+1].height = 75
                                        
                                        logger.info(f"已將照片插入到Excel第{row_idx+1}行")
                                except Exception as img_error:
                                    logger.error(f"處理圖片內存操作時發生錯誤: {img_error}")
                            else:
                                logger.warning(f"照片不存在或太小: {photo_path}")
                        except Exception as e:
                            logger.error(f"插入照片到Excel時發生錯誤: {str(e)}")
                    
                    # 保存修改後的Excel
                    wb.save(excel_path)
                    logger.info(f"已保存含照片的Excel文件: {excel_path}")
                    
                except ImportError as ie:
                    logger.warning(f"未安裝必要的庫，無法插入圖片: {ie}")
                    # 簡單保存Excel，不包含圖片
                    df.to_excel(excel_path, index=False)
                    logger.info(f"已保存基本Excel: {excel_path}")
                except Exception as excel_error:
                    logger.error(f"生成Excel時發生錯誤: {excel_error}")
                    # 出錯時簡單保存
                    try:
                        # 再次嘗試清理數據
                        for column in df.columns:
                            if df[column].dtype == 'object':
                                df[column] = df[column].apply(lambda x: clean_text_for_excel(x) if isinstance(x, str) else x)
                        df.to_excel(excel_path, index=False)
                        logger.info(f"已保存基本Excel: {excel_path}")
                    except Exception as final_error:
                        logger.error(f"最終保存嘗試失敗: {final_error}")
                        # 保存純JSON格式數據
                        json_path = os.path.join(self.config.output_dir, f"履歷資料_共{current_page}頁_緊急備份_{int(time.time())}.json")
                        with open(json_path, 'w', encoding='utf-8') as f:
                            # 先清理數據避免JSON序列化失敗
                            clean_data = []
                            for card in all_resume_cards:
                                clean_card = {}
                                for key, value in card.items():
                                    if isinstance(value, str):
                                        clean_card[key] = clean_text_for_excel(value)
                                    else:
                                        clean_card[key] = value
                                clean_data.append(clean_card)
                            
                            json.dump(clean_data, f, ensure_ascii=False, indent=2)
                
                # 保存至JSON
                json_path = os.path.join(self.config.output_dir, f"履歷資料_共{current_page}頁_{int(time.time())}.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    resume_dict = df.to_dict(orient='records')
                    json.dump(resume_dict, f, ensure_ascii=False, indent=2)
                logger.info(f"已保存履歷資料至JSON: {json_path}")
                
                return all_resume_cards
            else:
                logger.warning("未找到任何履歷卡片")
                return []
                
        except Exception as e:
            logger.error(f"提取履歷卡片時發生異常: {e}")
            return []
    
    def extract_info_from_text(self, text):
        """使用直接字串比對方式從文本中提取求職者信息，處理連續字串"""
        # 初始化結果字典
        info = {
            'name': None,
            'age': None,
            'gender': None,
            'code': None,
            'update_date': None,
            'work_location': None,
            'living_location': None,
            'education': None,
            'desired_job': None,
            'experience_years': None,
            'work_history': None
        }
        
        # 將所有文本合併成一個字串
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        full_text = ' '.join(lines)
        logger.debug(f"處理文本長度: {len(full_text)}")
        
        # 提取姓名 - 直接提取<a class="name word-break-all">和</a>之間的字串
        if 'class="name word-break-all' in full_text:
            try:
                # 使用正則表達式精確匹配class="name word-break-all"標籤中的內容
                name_match = re.search(r'class="name word-break-all[^"]*"[^>]*>([^<]+)</a>', full_text)
                if name_match:
                    name_text = name_match.group(1).strip()
                    # 確認提取的文本不是"更新日"或其他非名字內容
                    if name_text and not any(keyword in name_text for keyword in ['更新日', '代碼', '希望工作地', '居住地']):
                        info['name'] = clean_text_for_excel(name_text)
                    else:
                        logger.warning(f"提取到的名字可能有誤: {name_text}，嘗試其他方法")
                else:
                    # 如果正則匹配失敗，嘗試更精確的字符串查找
                    # 找到首個 class="name word-break-all 的位置
                    name_tag_start = full_text.find('class="name word-break-all')
                    if name_tag_start != -1:
                        # 找到標籤結束的 > 位置
                        close_tag_idx = full_text.find('>', name_tag_start)
                        if close_tag_idx != -1:
                            # 找到下一個 </a> 的位置
                            end_tag_idx = full_text.find('</a>', close_tag_idx)
                            if end_tag_idx != -1 and end_tag_idx > close_tag_idx:
                                # 提取 > 和 </a> 之間的內容
                                name_text = full_text[close_tag_idx+1:end_tag_idx].strip()
                                # 驗證提取的名字
                                if name_text and not any(keyword in name_text for keyword in ['更新日', '代碼', '希望工作地', '居住地']):
                                    info['name'] = clean_text_for_excel(name_text)
            except Exception as e:
                logger.debug(f"從HTML標籤提取名字時出錯: {e}")
        else:
            # 如果上述方法沒有提取到有效名字，嘗試從卡片文本的開頭部分提取名字
            if not info['name'] or any(keyword in info['name'] for keyword in ['更新日', '代碼', '希望工作地', '居住地']):
                # 移除HTML標籤，提取純文本
                clean_text = re.sub(r'<[^>]+>', ' ', full_text).strip()
                # 從純文本中提取名字，假設名字在開頭，後面跟著年齡或其他信息
                name_age_match = re.search(r'^\s*(\S+(?:\s+\S+){0,3})\s+(\d{1,2})歲', clean_text)
                if name_age_match:
                    potential_name = name_age_match.group(1).strip()
                    if potential_name and not any(keyword in potential_name for keyword in ['更新日', '代碼', '希望工作地', '居住地']):
                        info['name'] = clean_text_for_excel(potential_name)
        
        # 提取年齡
        age_match = re.search(r'(\d{1,2})歲', full_text)
        if age_match:
            info['age'] = clean_text_for_excel(age_match.group(1) + '歲')
        
        # 提取性別
        if '男' in full_text[:30]:
            info['gender'] = '男'
        elif '女' in full_text[:30]:
            info['gender'] = '女'
        
        # 提取代碼和更新日期
        code_match = re.search(r'代碼：(\d+)', full_text)
        if code_match:
            info['code'] = clean_text_for_excel(code_match.group(1))
        
        update_match = re.search(r'更新日：(\d{4}/\d{1,2}/\d{1,2})', full_text)
        if update_match:
            info['update_date'] = clean_text_for_excel(update_match.group(1))
        
        # 處理連續欄位的情況
        
        # 1. 找出所有關鍵字的位置
        work_loc_idx = full_text.find('希望工作地')
        living_loc_idx = full_text.find('居住地')
        
        # 尋找學歷關鍵字（碩士、大學、博士等）
        edu_keywords = ['碩士', '大學', '博士', '學士', '高中', '二技', '四技', '高職', '專科', '二專', '三專', '四專', '五專', '五技', '四技', '三技', '二技', '一技', '國中', '學士後', '學士後二技', '學士後四技', '學士後五技', '學士後四技', '學士後三技', '學士後二技', '學士後一技']
        edu_idx = -1
        edu_keyword = None
        
        for keyword in edu_keywords:
            idx = full_text.find(keyword)
            if idx != -1:
                edu_idx = idx
                edu_keyword = keyword
                break
        
        desired_job_idx = full_text.find('希望職稱')
        exp_idx = full_text.find('工作經驗')
        
        # 2. 提取希望工作地
        if work_loc_idx != -1 and living_loc_idx != -1 and work_loc_idx < living_loc_idx:
            colon_idx = full_text.find(':', work_loc_idx)
            if colon_idx == -1:
                colon_idx = full_text.find('：', work_loc_idx)
            
            if colon_idx != -1:
                # 希望工作地的內容從冒號之後到居住地之前
                work_location = full_text[colon_idx+1:living_loc_idx].strip()
                info['work_location'] = clean_text_for_excel(work_location)
        
        # 3. 提取居住地
        if living_loc_idx != -1 and edu_idx != -1 and living_loc_idx < edu_idx:
            colon_idx = full_text.find(':', living_loc_idx)
            if colon_idx == -1:
                colon_idx = full_text.find('：', living_loc_idx)
            
            if colon_idx != -1:
                # 居住地的內容從冒號之後到學歷之前
                living_location = full_text[colon_idx+1:edu_idx].strip()
                info['living_location'] = clean_text_for_excel(living_location)
        
        # 4. 提取學歷
        if edu_idx != -1 and desired_job_idx != -1 and edu_idx < desired_job_idx:
            # 學歷的內容從學歷關鍵字開始到希望職稱之前
            education = full_text[edu_idx:desired_job_idx].strip()
            info['education'] = clean_text_for_excel(education)
        
        # 5. 提取希望職稱
        if desired_job_idx != -1 and desired_job_idx < exp_idx:
            colon_idx = full_text.find(':', desired_job_idx)
            if colon_idx == -1:
                colon_idx = full_text.find('：', desired_job_idx)
            
            if colon_idx != -1:
                # 先提取冒號後的內容
                job_content = full_text[colon_idx+1:]
                
                # 尋找數字開頭的工作年資（如「1~2年工作經驗」、「3年經驗」等）
                year_exp_match = re.search(r'[1-9][\d~]*年', job_content)
                
                if year_exp_match:
                    # 如果找到數字年資，則以此為界限
                    year_exp_pos = year_exp_match.start()
                    desired_job = job_content[:year_exp_pos].strip()
                else:
                    # 如果沒找到數字年資，則仍以「工作經驗」為界限
                    exp_pos = job_content.find('工作經驗')
                    if exp_pos != -1:
                        desired_job = job_content[:exp_pos].strip()
                    else:
                        # 如果都沒找到，就取整段內容
                        desired_job = job_content.strip()
                
                info['desired_job'] = clean_text_for_excel(desired_job)

            # 5.5 提取工作時長（希望職稱與工作經驗之間的內容）
            if desired_job_idx != -1 and exp_idx != -1 and desired_job_idx < exp_idx:
                # 找到希望職稱後的冒號
                colon_idx = full_text.find(':', desired_job_idx)
                if colon_idx == -1:
                    colon_idx = full_text.find('：', desired_job_idx)
                
                if colon_idx != -1 and info['desired_job']:
                    # 計算工作時長的起始位置（希望職稱內容之後）
                    start_pos = colon_idx + 2 + len(info['desired_job'])
                    # 工作時長為希望職稱與工作經驗之間的內容
                    experience_duration = full_text[start_pos:exp_idx].strip()
                    
                    # 設定experience_years為這段內容
                    if experience_duration:
                        info['experience_years'] = clean_text_for_excel(experience_duration)
        
        # 6. 提取工作經驗
        if exp_idx != -1:
            # 提取年資信息（例如「1~2年工作經驗」）
            exp_pattern = r'((?:[1-9][\d]*~[1-9][\d]*年|[1-9][\d]*年以[上下]|[<>][1-9][\d]*年|[1-9][\d]*年))工作經驗'
            exp_match = re.search(exp_pattern, full_text[exp_idx:])
            if exp_match:
                full_exp = exp_match.group(0)  # 完整匹配，包含「工作經驗」
                years_only = exp_match.group(1)  # 只有年資部分，如「1~2年」
                info['experience_years'] = clean_text_for_excel(years_only)
                
                # 計算工作經歷開始的位置
                exp_full_pos = full_text.find(full_exp, exp_idx)
                work_history_start_idx = exp_full_pos + len(full_exp)
            else:
                # 如果找不到特定格式，則從「工作經驗」之後開始
                work_history_start_idx = exp_idx + len('工作經驗')
            
            # 提取工作經歷
            work_history_content = full_text[work_history_start_idx:]
            
            # 找到邀約、儲存等操作按鈕前的文本作為工作經歷
            action_keywords = ['邀約', '儲存', '轉寄', '備註']
            end_idx = len(work_history_content)
            
            for keyword in action_keywords:
                idx = work_history_content.find(keyword)
                if idx != -1 and idx < end_idx:
                    end_idx = idx
            
            work_history = work_history_content[:end_idx].strip()
            info['work_history'] = clean_text_for_excel(work_history)
        
        return info
    
    async def extract_photo_url(self, card):
        """從卡片中提取大頭照URL"""
        # 首先嘗試提取104特定格式的頭像
        try:
            # 直接尋找104 VIP網站特定格式的頭像圖片
            img = await card.query_selector('img[src*="webHeadShot"]')
            if img:
                src = await img.get_attribute('src')
                if src:
                    logger.info(f"找到104特定格式頭像: {src}")
                    return src
        except Exception as e:
            logger.debug(f"提取104特定頭像時出錯: {e}")
        
        # 如果找不到特定格式，嘗試其他選擇器
        photo_selectors = [
            'img[src*="headShot"]', 
            'img[src*="photo"]', 
            'img.avatar', 
            'img.profile-photo', 
            '.photo img', 
            '.avatar img',
            'img[title]',  # 104經常在title屬性放人名
            '.BaseAvatar img',
            'img[alt*="照片"]'
        ]
        
        for selector in photo_selectors:
            try:
                imgs = await card.query_selector_all(selector)
                for img in imgs:
                    src = await img.get_attribute('src')
                    if src and ('webHeadShot' in src or 'headShot' in src or 'photo' in src or 'avatar' in src):
                        # 如果是相對URL，轉換為絕對URL
                        if src.startswith('/'):
                            src = f"https://vip.104.com.tw{src}"
                        logger.info(f"找到候選頭像: {src}")
                        return src
            except Exception as e:
                logger.debug(f"使用選擇器 '{selector}' 提取頭像URL失敗: {e}")
        
        return None
    
    async def download_photo(self, url, save_path):
        """簡化的照片下載方法"""
        try:
            # 確保URL和保存路徑經過清理
            url = clean_text_for_excel(url)
            save_path_dir = os.path.dirname(save_path)
            save_path_filename = os.path.basename(save_path)
            # 確保文件名安全
            save_path_filename = self.sanitize_filename(save_path_filename)
            # 重新組合路徑
            save_path = os.path.join(save_path_dir, save_path_filename)
            
            logger.info(f"開始下載大頭照: {url}")
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            # 使用curl命令下載照片
            cookies = await self.browser.cookies()
            cookie_str = '; '.join([f"{c['name']}={c['value']}" for c in cookies 
                                    if 'vip.104.com.tw' in c.get('domain', '') or 
                                    'asset.vip.104.com.tw' in c.get('domain', '')])
            
            # 創建臨時文件
            with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as temp:
                temp_path = temp.name
            
            # 構建curl命令 - 關閉SSL驗證
            curl_cmd = [
                'curl', '-k', '--retry', '3', '--retry-delay', '2',
                '-L', url,
                '-H', f'Cookie: {cookie_str}',
                '-H', 'Referer: https://vip.104.com.tw/',
                '-H', 'User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
                '-o', temp_path
            ]
            
            # 執行命令
            proc = await asyncio.create_subprocess_exec(
                *curl_cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            # 等待完成
            _, stderr = await proc.communicate()
            
            if proc.returncode == 0 and os.path.exists(temp_path):
                file_size = os.path.getsize(temp_path)
                if file_size > 100:
                    shutil.move(temp_path, save_path)
                    logger.info(f"大頭照下載成功: {save_path}")
                    return True
                else:
                    logger.warning(f"下載的照片太小: {file_size} bytes")
                    os.unlink(temp_path)
            else:
                error = stderr.decode('utf-8', errors='ignore') if stderr else "未知錯誤"
                logger.warning(f"curl下載失敗: {error}")
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            
            return False
            
        except Exception as e:
            logger.error(f"下載照片過程中發生錯誤: {str(e)}")
            return False
    
    def sanitize_filename(self, filename):
        """確保文件名有效（移除不允許的字符）"""
        # 替換不允許用作文件名的字符
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename
    
    async def run(self):
        """執行完整爬蟲流程，使用持久化上下文保存狀態"""
        try:
            # 初始化瀏覽器
            await self.initialize()
            
            # 登入
            login_success = await self.login()
            if not login_success:
                logger.error("登入失敗，無法繼續爬蟲")
                return False
            
            # 搜尋
            if self.config.search_keyword:
                search_success = await self.search()
                if not search_success:
                    logger.error("搜尋失敗，無法提取結果")
                    return False
                
                # 提取結果
                results = await self.extract_results()
                return results
            else:
                logger.info("未設定搜尋關鍵字，跳過搜尋步驟")
                return True
            
        except Exception as e:
            logger.error(f"爬蟲執行過程發生異常: {e}")
            return False
    
    async def close(self):
        """關閉瀏覽器"""
        if self.browser:
            await self.browser.close()
            logger.info("瀏覽器已關閉")
        # 在持久化上下文中，context 和 browser 是同一個對象，不需要額外關閉

    async def check_if_logged_in(self):
        """檢查是否已經登入VIP系統"""
        try:
            # 方法1：檢查URL是否包含index/index
            if "/index/index" in self.page.url:
                return True
            
            # 方法2：檢查是否存在退出按鈕
            logout_selectors = [
                'a:text("登出")', 
                'button:text("登出")',
                '[href*="logout"]'
            ]
            
            for selector in logout_selectors:
                try:
                    logout_element = await self.page.query_selector(selector)
                    if logout_element:
                        return True
                except:
                    continue
            
            # 方法3：檢查頁面是否包含用戶名稱
            try:
                # 嘗試提取用戶名或公司名相關元素
                username_content = await self.page.evaluate('''() => {
                    // 嘗試各種可能包含用戶名的元素
                    const userElements = document.querySelectorAll('.user-name, .username, .account-name');
                    for (const el of userElements) {
                        if (el.textContent.trim()) {
                            return el.textContent.trim();
                        }
                    }
                    return '';
                }''')
                
                if username_content and len(username_content) > 1:
                    logger.info(f"檢測到用戶名: {username_content}")
                    return True
            except Exception as e:
                logger.debug(f"檢查用戶名時出錯: {e}")
            
            return False
        except Exception as e:
            logger.error(f"檢查登入狀態時出錯: {e}")
            return False

    async def go_to_next_page(self):
        """前往下一頁"""
        try:
            logger.info("嘗試前往下一頁")
            
            # 嘗試多種下一頁按鈕選擇器
            next_page_selectors = [
                '.paging-container a.pager_next',
                'a.pager_next',
                'a.next-page',
                'a:has-text("下一頁")',
                'button:has-text("下一頁")',
                '.pagination .next',
                '.pagination-next',
                '.pagination li:last-child a:not(.disabled)',
                '[aria-label="Next page"]',
                '[data-qa-id="nextPage"]',
                '.pager-next',
                'a[rel="next"]'
            ]
            
            for selector in next_page_selectors:
                try:
                    next_button = await self.page.query_selector(selector)
                    if next_button and await next_button.is_visible():
                        # 檢查按鈕是否被禁用
                        is_disabled = await self.page.evaluate(
                            '(element) => element.disabled || element.classList.contains("disabled") || element.getAttribute("aria-disabled") === "true"', 
                            next_button
                        )
                        
                        if not is_disabled:
                            # 捲動到下一頁按鈕位置
                            await next_button.scroll_into_view_if_needed()
                            await asyncio.sleep(1)
                            
                            # 點擊下一頁按鈕
                            await next_button.click()
                            logger.info("已點擊下一頁按鈕")
                            
                            # 等待頁面加載
                            await self.page.wait_for_load_state('networkidle', timeout=15000)
                            await asyncio.sleep(3)  # 等待額外時間確保頁面載入完成
                            
                            # 檢查頁面是否有變化
                            current_url = self.page.url
                            if "page=" in current_url or "p=" in current_url:
                                logger.info(f"成功進入下一頁：{current_url}")
                                return True
                            else:
                                logger.info("頁面URL未顯示變化，檢查頁面內容是否已更新")
                                return True
                except Exception as e:
                    logger.debug(f"使用選擇器 '{selector}' 點擊下一頁按鈕失敗: {e}")
                    continue
            
            # 如果所有選擇器都失敗，嘗試通過URL參數直接進入下一頁
            try:
                current_url = self.page.url
                page_match = re.search(r'[?&](page|p)=(\d+)', current_url)
                
                if page_match:
                    # 找到頁數參數，如 page=1 或 p=1
                    param_name = page_match.group(1)  # page 或 p
                    current_page = int(page_match.group(2))
                    next_page = current_page + 1
                    
                    # 替換頁數
                    next_url = re.sub(
                        f'{param_name}={current_page}', 
                        f'{param_name}={next_page}', 
                        current_url
                    )
                    
                    logger.info(f"通過URL參數進入下一頁: {next_url}")
                    await self.page.goto(next_url, timeout=30000)
                    await self.page.wait_for_load_state('networkidle', timeout=15000)
                    await asyncio.sleep(3)
                    return True
                else:
                    # URL中沒有頁數參數，追加頁數參數
                    separator = "&" if "?" in current_url else "?"
                    next_url = f"{current_url}{separator}page=2"
                    
                    logger.info(f"通過增加頁數參數進入第二頁: {next_url}")
                    await self.page.goto(next_url, timeout=30000)
                    await self.page.wait_for_load_state('networkidle', timeout=15000)
                    await asyncio.sleep(3)
                    return True
                    
            except Exception as e:
                logger.error(f"通過URL參數進入下一頁失敗: {e}")
            
            logger.warning("嘗試所有方法進入下一頁失敗，可能已經是最後一頁")
            return False
            
        except Exception as e:
            logger.error(f"前往下一頁過程中發生錯誤: {e}")
            return False

async def main():
    """主程序"""
    print("=== 104人力銀行求職者爬蟲 ===")
    print("注意：使用本工具需遵守104相關使用條款及個人資料保護法")
    print("      僅供學習研究使用，請勿用於商業或非法用途\n")
    
    # 檢查是否有已儲存的使用者資訊
    config_file = "user_config.json"
    saved_config = {}
    
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                saved_config = json.load(f)
            print("找到已儲存的帳號資訊")
        except:
            pass
    
    # 詢問是否使用已儲存的帳號
    use_saved = False
    if saved_config.get("username"):
        use_saved_input = input(f"是否使用已儲存的帳號 ({saved_config.get('username')})？(y/n): ")
        use_saved = use_saved_input.lower() == 'y'
    
    if use_saved and saved_config.get("username") and saved_config.get("password"):
        username = saved_config.get("username")
        password = saved_config.get("password")
        print(f"使用已儲存的帳號: {username}")
    else:
        username = input("請輸入104企業會員帳號: ")
        password = input("請輸入104企業會員密碼: ")
        
        # 詢問是否記住帳號密碼
        save_account = input("是否記住帳號密碼？(y/n): ")
        if save_account.lower() == 'y':
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump({"username": username, "password": password}, f)
            print("已儲存帳號資訊")
    
    keyword = input("請輸入搜索關鍵詞 (直接按Enter搜索全部): ")
    
    # 新增頁數輸入
    page_limit_input = input("請輸入要爬取的頁數 (預設為1頁): ")
    page_limit = 1
    try:
        if page_limit_input.strip():
            page_limit = int(page_limit_input)
            if page_limit < 1:
                print("頁數必須大於0，使用預設值1頁")
                page_limit = 1
    except ValueError:
        print("輸入的頁數無效，使用預設值1頁")
    
    # 創建設定
    config = ResumeScraperConfig(
        username=username,
        password=password,
        search_keyword=keyword,
        page_limit=page_limit
    )
    
    # 創建爬蟲實例
    scraper = ResumeScraper(config)
    
    try:
        # 執行爬蟲
        results = await scraper.run()
        
        if results and isinstance(results, list):
            print(f"爬蟲完成，共獲取 {len(results)} 份履歷")
            print(f"結果已保存至目錄: {config.output_dir}")
        elif results is True:
            print("爬蟲流程已完成")
        else:
            print("爬蟲未能獲取有效結果")
    
    except Exception as e:
        logger.error(f"程序執行時出錯: {e}")
        print(f"程序執行時出錯: {e}")
    
    finally:
        # 關閉瀏覽器
        await scraper.close()

if __name__ == "__main__":
    asyncio.run(main())